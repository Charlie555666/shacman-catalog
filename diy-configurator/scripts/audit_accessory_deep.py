#!/usr/bin/env python3
"""
Deep audit: Compare ALL accessory swap entries against Excel source.
Special focus on 轮胎(tires), 空滤(air filter), 油箱(fuel tank), etc.
"""
import json, re, sys
from collections import defaultdict
import openpyxl

BASE = "diy-configurator/data"
EXCEL_PATH = "附件2-2026年出口产品换装价格表.xlsx"

# ─── Load JSON ───
with open(f"{BASE}/accessory_swaps.json", "r", encoding="utf-8") as f:
    swaps = json.load(f)

cats = swaps.get("swap_categories", {})

# ─── Load Excel ───
wb = openpyxl.load_workbook(EXCEL_PATH, data_only=True)
ws = wb["柴油车"]

excel_entries = []
for row in ws.iter_rows(min_row=3, max_row=ws.max_row, values_only=True):
    seq, category, content, unit, price, note = (
        row[0], row[1], row[2], row[3], row[4], 
        row[5] if len(row) > 5 else None
    )
    seq = int(seq) if isinstance(seq, (int, float)) and seq == seq else None
    category = str(category).strip() if category else ""
    content = str(content).strip() if content else ""
    price = float(price) if isinstance(price, (int, float)) and price == price else None
    note = str(note).strip() if note else ""
    
    if content:
        excel_entries.append({
            "seq": seq,
            "category": category,
            "content": content,
            "price": price,
            "note": note
        })

print(f"Excel entries loaded: {len(excel_entries)}")
print(f"JSON categories: {len(cats)}")
print()

# ─── Helper: Normalize text for matching ───
def norm(s):
    """Remove spaces, standardize separators."""
    s = s.replace(" ", "").replace("\u3000", "").replace("）", ")").replace("（", "(")
    s = s.replace("→", "→").replace("->", "→")
    # Don't replace 换/换装/选装/指定 here — handle in matching
    return s

def fuzzy_match(json_entry, excel_entries):
    """Try to find matching Excel entry for a JSON swap entry.
    Returns (match, confidence_level, price_diff)
    """
    source = norm(json_entry["source"])
    target = norm(json_entry["target"])
    json_price = json_entry["price"]
    is_deduction = json_entry.get("is_deduction", False)
    
    best = None
    
    for ex in excel_entries:
        ex_content = norm(ex["content"])
        
        # Build many candidate match patterns
        # Pattern types:
        # 1. 换装: source + 换装 + target ("普通空滤换装中置空滤器")
        # 2. 换: source + 换 + target
        # 3. 指定/选装: source + 指定/选装/换装 + target ("12.00R20指定三角品牌")
        # 4. Reverse: target回source (降配)
        
        match_score = 0
        
        # Exact patterns
        patterns_exact = [
            f"{source}换装{target}",
            f"{source}换{target}",
            f"{source}指定{target}",
            f"{source}选装{target}",
            f"{source}→{target}",
        ]
        
        for p in patterns_exact:
            if ex_content == p:
                match_score = 5
                break
            elif p in ex_content:
                match_score = max(match_score, 3)
        
        # Partial: source and target both appear (but NOT in opposite direction)
        if match_score < 3:
            # Check if both source and target appear in content
            has_source = source in ex_content
            has_target = target in ex_content
            
            # Avoid matching if it looks like target→source (reverse)
            reverse_patterns = [f"{target}换装{source}", f"{target}换{source}", 
                              f"{target}指定{source}", f"{target}选装{source}"]
            
            is_reverse = any(p in ex_content for p in reverse_patterns)
            
            if has_source and has_target and not is_reverse:
                match_score = 2
            elif has_source and has_target:
                match_score = 0  # It's a reverse match, skip
            elif has_source:
                # Source matches but target doesn't — could be a tire brand where
                # target is the brand, like "三角" appearing separately
                if len(target) > 1 and target in ex_content:
                    match_score = 2
        
        if match_score > 0:
            ex_price = ex["price"]
            json_effective = json_price * (-1 if is_deduction else 1)
            
            if ex_price is not None:
                price_diff = json_effective - ex_price
                is_match = abs(price_diff) < 10
            else:
                price_diff = None
                is_match = None
            
            if best is None or match_score > best["score"]:
                best = {
                    "score": match_score,
                    "excel": ex,
                    "price_diff": price_diff,
                    "is_match": is_match
                }
    
    return best

# ─── Audit ───
results = {"matched_ok": 0, "matched_mismatch": 0, "unmatched": 0, "price_na": 0}
details = {"mismatches": [], "unmatched": [], "by_category": defaultdict(lambda: {"ok": 0, "mismatch": 0, "unmatched": 0, "na": 0})}

# Also scan for items in Excel not in JSON (coverage check)
excel_matched = [False] * len(excel_entries)

total_checked = 0

for cat_name, cat_data in sorted(cats.items()):
    opts = cat_data.get("options", {})
    for opt_name, targets in opts.items():
        for t in targets:
            total_checked += 1
            match = fuzzy_match(t, excel_entries)
            
            if match and match["is_match"] is True:
                results["matched_ok"] += 1
                details["by_category"][cat_name]["ok"] += 1
                # Mark excel entry as matched
                for i, ex in enumerate(excel_entries):
                    if ex is match["excel"]:
                        excel_matched[i] = True
            elif match and match["is_match"] is False:
                results["matched_mismatch"] += 1
                details["by_category"][cat_name]["mismatch"] += 1
                details["mismatches"].append({
                    "category": cat_name,
                    "source": t["source"],
                    "target": t["target"],
                    "json_price": t["price"],
                    "is_deduction": t.get("is_deduction", False),
                    "excel_price": match["excel"]["price"],
                    "excel_content": match["excel"]["content"],
                    "price_diff": match["price_diff"],
                    "excel_note": match["excel"]["note"]
                })
            elif match and match["price_diff"] is None:
                results["price_na"] += 1
                details["by_category"][cat_name]["na"] += 1
            else:
                results["unmatched"] += 1
                details["by_category"][cat_name]["unmatched"] += 1
                details["unmatched"].append({
                    "category": cat_name,
                    "source": t["source"],
                    "target": t["target"],
                    "json_price": t["price"],
                    "is_deduction": t.get("is_deduction", False)
                })

# ─── Report ───
print("=" * 80)
print("ACCESSORY SWAPS DEEP AUDIT REPORT")
print("=" * 80)
print(f"\nTotal JSON entries checked: {total_checked}")
print(f"  ✅ Matched & correct: {results['matched_ok']}")
print(f"  ❌ Matched & MISMATCH: {results['matched_mismatch']}")
print(f"  ⚠️  Unmatched (no Excel ref): {results['unmatched']}")
print(f"  ⬜ Price N/A in Excel: {results['price_na']}")

print(f"\n{'─' * 80}")
print(f"{'Category':<20s} {'✅OK':>6s} {'❌Mismatch':>10s} {'⚠️Unmatched':>12s} {'⬜N/A':>6s} {'Total':>6s}")
print(f"{'─' * 80}")

for cat_name in sorted(details["by_category"].keys()):
    d = details["by_category"][cat_name]
    total = d["ok"] + d["mismatch"] + d["unmatched"] + d["na"]
    print(f"{cat_name:<20s} {d['ok']:>6d} {d['mismatch']:>10d} {d['unmatched']:>12d} {d['na']:>6d} {total:>6d}")

# ─── Mismatch details ───
if details["mismatches"]:
    print(f"\n{'=' * 80}")
    print(f"❌ PRICE MISMATCHES ({len(details['mismatches'])} found):")
    print(f"{'=' * 80}")
    for m in details["mismatches"]:
        effective = m["json_price"] * (-1 if m["is_deduction"] else 1)
        print(f"\n  [{m['category']}] {m['source']} → {m['target']}")
        print(f"    JSON: ¥{effective} (raw={m['json_price']}, is_deduction={m['is_deduction']})")
        print(f"    Excel: ¥{m['excel_price']} ({m['excel_content']})")
        print(f"    Diff: ¥{m['price_diff']}")
        if m["excel_note"]:
            print(f"    Excel note: {m['excel_note']}")

# ─── Coverage: Excel items NOT in JSON ───
print(f"\n{'=' * 80}")
print(f"EXCEL COVERAGE CHECK — Items in Excel but NOT in JSON:")
print(f"{'=' * 80}")
missed = []
for i, matched in enumerate(excel_matched):
    if not matched and excel_entries[i]["seq"]:
        ex = excel_entries[i]
        missed.append(ex)

if missed:
    print(f"\n{len(missed)} Excel items not found in JSON:")
    for ex in missed:
        print(f"  #{ex['seq']:>3} | [{ex['category']}] {ex['content'][:80]} | ¥{ex['price']}")
else:
    print("  All Excel items appear to be covered in JSON (or are add-on items without swap pairs).")

# ─── Special: Spot-check air filter prices ───
print(f"\n{'=' * 80}")
print("SPOT-CHECK: 空滤 (Air Filter) — All entries")
print(f"{'=' * 80}")
for cat_name in ["其他附件"]:
    cat_data = cats.get(cat_name, {})
    opts = cat_data.get("options", {})
    for opt_name, targets in opts.items():
        if "滤" in opt_name:
            print(f"\n  Source: {opt_name}")
            for t in targets:
                effective = t["price"] * (-1 if t.get("is_deduction") else 1)
                # Find excel match
                match = fuzzy_match(t, excel_entries)
                ex_str = ""
                if match:
                    ex_str = f"Excel: ¥{match['excel']['price']} | {'✅' if match['is_match'] else '❌ DIFF=' + str(match['price_diff'])}"
                else:
                    ex_str = "⚠️ No Excel match"
                print(f"    → {t['target']}: ¥{effective} (raw={t['price']}, deduct={t.get('is_deduction')}) | {ex_str}")

# ─── Special: Spot-check tires ───
print(f"\n{'=' * 80}")
print("SPOT-CHECK: 轮胎品牌 (Tire Brands) — First 10 entries")
print(f"{'=' * 80}")
cat_data = cats.get("轮胎品牌", {})
opts = cat_data.get("options", {})
count = 0
for opt_name, targets in opts.items():
    if count >= 10:
        break
    for t in targets:
        if count >= 10:
            break
        effective = t["price"] * (-1 if t.get("is_deduction") else 1)
        match = fuzzy_match(t, excel_entries)
        ex_str = ""
        if match:
            match_sym = "✅" if match["is_match"] else "❌"
            ex_str = f"Excel: ¥{match['excel']['price']} | {match_sym}"
        else:
            ex_str = "⚠️ No Excel match"
        print(f"  {t['source'][:25]} → {t['target'][:25]}: ¥{effective} | {ex_str}")
        count += 1

# Final summary
print(f"\n{'=' * 80}")
if results["matched_mismatch"] == 0:
    print("✅ AUDIT PASSED — All matched entries have correct prices.")
else:
    print(f"❌ AUDIT FAILED — {results['matched_mismatch']} price mismatches found!")
print(f"Total unmatched (likely auto-generated reverse entries): {results['unmatched']}")
print(f"{'=' * 80}")
