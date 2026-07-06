"""逐条对比 engine_swaps.json 与源 Excel"""
import json, os, sys
from pathlib import Path

DATA_DIR = Path(__file__).parent.parent / 'data'

with open(DATA_DIR / 'engine_swaps.json', 'r', encoding='utf-8') as f:
    engine_data = json.load(f)

import openpyxl
wb = openpyxl.load_workbook(r'C:\Users\Administrator\Desktop\陕汽报价\V1-发动机换选装价格（2）.xlsx', data_only=True)
ws = wb['Sheet1']

# Build Excel map
excel_rules = {}
for row in ws.iter_rows(min_row=2, max_row=ws.max_row, values_only=True):
    if not row[0]:
        continue
    std = str(row[1]).strip() if row[1] else ''
    swp = str(row[2]).strip() if row[2] else ''
    try:
        price_wan = float(row[3])
    except:
        continue
    price = int(price_wan * 10000)  # 万→元
    key = (std, swp)
    excel_rules[key] = {'price_wan': price_wan, 'price': price, 'row': row[0]}

# Compare JSON rules to Excel
errors = []
matched = 0
unknown = 0
dupes = set()

for i, r in enumerate(engine_data['rules']):
    key = (r['standard'], r['swap'])
    
    if key in dupes:
        continue
    dupes.add(key)
    
    if key in excel_rules:
        excel = excel_rules[key]
        json_price = r['price_change']
        excel_price = excel['price']
        
        if json_price != excel_price:
            errors.append({
                'pair': f"{key[0]} → {key[1]}",
                'excel_row': excel['row'],
                'json_price': json_price,
                'excel_price': excel_price,
                'excel_wan': excel['price_wan']
            })
        else:
            matched += 1
    else:
        # Check if reverse exists in Excel
        rev_key = (key[1], key[0])
        if rev_key in excel_rules:
            rev_excel = excel_rules[rev_key]
            expected_rev = round(abs(rev_excel['price']) * 0.8)
            if r['price_change'] < 0:
                expected_rev = -expected_rev
            actual = r['price_change']
            if actual != expected_rev:
                errors.append({
                    'pair': f"{key[0]} → {key[1]}",
                    'excel_row': f"(auto-gen from row {rev_excel['row']}: {rev_key[0]}→{rev_key[1]}={rev_excel['price']})",
                    'json_price': actual,
                    'excel_price': expected_rev,
                    'excel_wan': expected_rev/10000
                })
            else:
                matched += 1
        else:
            unknown += 1

print(f"逐条对比结果:")
print(f"  Excel原始规则: {len(excel_rules)}条")
print(f"  JSON rules: {len(engine_data['rules'])}条")
print(f"  ✅ 匹配: {matched}条")
print(f"  ❓ 无法匹配(可能为特殊规则): {unknown}条")
print(f"  ❌ 价格不匹配: {len(errors)}条")

if errors:
    print(f"\n{'='*60}")
    print("价格错误详情:")
    for e in errors:
        print(f"  {e['pair']}")
        print(f"    Excel: {e['excel_wan']}万 ({e['excel_price']}元) [行{e['excel_row']}]")
        print(f"    JSON:  {e['json_price']}元")
        diff = e['json_price'] - e['excel_price']
        if diff != 0:
            print(f"    ⚠️ 差异: {diff:+,}元")
        print()
else:
    print("\n🎉 所有价格与Excel完全一致!")
